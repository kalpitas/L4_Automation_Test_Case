from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.support.ui import Select
import time
from openpyxl import Workbook,load_workbook


def test_21cfr(chrome_browser):

    #chrome_browser.implicitly_wait(60)
    wb = load_workbook("ACG_Common_Workbook.xlsx")
    pass_comp = wb["Password_Complex"]
    reset_link = pass_comp.cell(row = 1, column = 2).value
    print(reset_link)

    start_row = 5
    last_row =  (pass_comp.max_row) +1

    chrome_browser.get(str(reset_link))

    timeout = 5
    try:
        subm = EC.element_to_be_clickable((By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[2]/div/div[6]/button[2]"))
        WebDriverWait(chrome_browser,timeout).until(subm)
    except TimeoutException:
        print("Submit button not visible")

    for i in range(start_row, last_row):

        test_case = pass_comp.cell(row=i, column=5).value

        print(f'======================{test_case}=========================')

        umail = pass_comp.cell(row = i,column = 1).value
        chrome_browser.find_element(By.NAME,"UserEmail").send_keys(str(umail))

        if test_case not in ("Validate a password with no characters"):

            upass = pass_comp.cell(row=i, column=2).value
            chrome_browser.find_element(By.NAME, "Password").send_keys(str(upass))

            ucpass = pass_comp.cell(row=i, column=3).value
            chrome_browser.find_element(By.NAME, "confirmPassword").send_keys(str(ucpass))

        chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[2]/div/div[6]/button[2]").click()

        try:
            loginbut = EC.element_to_be_clickable((By.XPATH,"/html/body/div[3]/div[3]/div/section/footer/button"))
            WebDriverWait(chrome_browser,timeout).until(loginbut)
            message = "Password got reset successfully"
        except TimeoutException:
            chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[2]/div/div[6]/button[2]").click()
            if test_case not in ("Validate a password reset with difference in password and confirm password tab"):
                message = chrome_browser.find_element(By.XPATH,"/html/body/div[1]/div/div/div[4]/div/div[2]/div/div[4]/div/div[2]").text
            else:
                message = chrome_browser.find_element(By.XPATH,"/html/body/div[1]/div/div/div[4]/div/div[2]/div/div[5]/div/div[2]").text

        pass_comp.cell(row = i, column = 4).value = message
        wb.save("ACG_Common_Workbook.xlsx")
        print(message)

        chrome_browser.refresh()



