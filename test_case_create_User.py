from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
import time
from openpyxl import Workbook,load_workbook
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from login_page_user import *

def test_create_user(chrome_browser):


    #chrome_browser.implicitly_wait(60)
    wb = load_workbook("ACG_Common_Workbook.xlsx")
    typeev = wb["URL_Login_cred_Tenant"]
    type_env = typeev.cell(row = 2, column = 4).value

    if type_env == "ACG":
        #fetch login_cred sheet
        url_login_cred = wb["URL_Login_cred_ACG"]

    if type_env == "Tenant":
        #fetch login_cred sheet
        url_login_cred = wb["URL_Login_cred_Tenant"]

    '''
    url = url_login_cred.cell(row = 2, column = 1).value
    print(url)
    chrome_browser.get(str(url))

    #if type_env == 'ACG':
    chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

    usern = url_login_cred.cell(row = 2, column = 2).value
    chrome_browser.find_element(By.NAME,'userName').send_keys(str(usern))

    passw = url_login_cred.cell(row = 2, column = 3).value
    chrome_browser.find_element(By.NAME,'password').send_keys(str(passw))

    chrome_browser.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[2]/div/button").click()
    '''

    login(chrome_browser)

    user = wb['Users']
    start_user = 2
    last_user = (user.max_row) +1


    if type_env == 'ACG':
        chrome_browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[2]/h2/button/span').click()
        chrome_browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[2]/div/div/a[1]').click()

    if type_env == 'Tenant':
        chrome_browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[1]/h2/button/span').click()
        chrome_browser.find_element(By.XPATH,'/html/body/div[1]/div/div/div[2]/div[1]/div/div[1]/div/div/a[1]').click()


    for i in range(start_user, last_user):

        test_case = user.cell(row=i, column=7).value

        print(f'======================{test_case}=========================')
        timeout = 10
        try:
            new_user_nxt = EC.element_to_be_clickable((By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div[1]/button[2]"))
            WebDriverWait(chrome_browser,timeout).until(new_user_nxt)
        except TimeoutException:
            print("Create user page_nxt: Timed out waiting for page to load")

        if test_case not in ("To create a user without username"):
            name = user.cell(row = i, column = 1).value
            chrome_browser.find_element(By.NAME,"userName").send_keys(str(name))

        if test_case not in ("To create a user without emailid"):
            email = user.cell(row = i, column = 2).value
            chrome_browser.find_element(By.NAME,"emailAddress").send_keys(str(email))

        mob = user.cell(row = i, column = 3).value
        chrome_browser.find_element(By.NAME,"phoneNumber").send_keys(str(mob))

        if test_case not in ("To create a user without selecting role"):
            # select by visible text
            role = user.cell(row = i, column = 4).value
            print(role)
            select = Select(chrome_browser.find_element(By.NAME,"roleId"))
            select.select_by_visible_text(str(role))

        if type_env == 'Tenant' and test_case not in ("To create a user without selecting location"):
            location = user.cell(row = i, column = 5).value
            print(location)
            select = Select(chrome_browser.find_element(By.NAME,"locationId"))
            select.select_by_visible_text(str(location))

        #click on next button
        chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[3]/div[1]/button[2]").click()
        print("clicked on next button")

        timeout = 6
        try:
            #click on submit
            new_user_sub = EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]"))
            WebDriverWait(chrome_browser, timeout).until(new_user_sub)
            chrome_browser.find_element(By.XPATH, "//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]").click()
            print("clicked on submit button")
            try:
                # click on new user
                new_user = EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b"))
                WebDriverWait(chrome_browser, timeout).until(new_user)
                #successfull message
                message = chrome_browser.find_element(By.XPATH,"/html/body/div[2]/div[4]/div/div/div/div/div[2]").text
                print(message)
            except TimeoutException:
                print("Nested Final exception")
                chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[3]/div[1]/button[2]").click()
                chrome_browser.find_element(By.XPATH, "//*[@id='root']/div/div/div/div[2]/div/div[1]/button[2]").click()
                message = chrome_browser.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/div/div/div[2]").text
                print(message)
                chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div[1]/button[1]").click()
        except TimeoutException:
            print("Final exception")
            chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div/div[2]/div/div[3]/div[1]/button[2]").click()
            message = chrome_browser.find_element(By.XPATH, "/html/body/div[2]/div[4]/div/div/div/div/div[2]").text
            print(message)
            chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div[3]/div[1]/button[1]").click()

        user.cell(row=i, column=6).value = message
        wb.save("ACG_Common_Workbook.xlsx")

        time.sleep(5)

        chrome_browser.find_element(By.XPATH,"//*[@id='root']/div/div/div[2]/div[2]/div/div/div[2]/b").click()
        print("clicked on new user button")
