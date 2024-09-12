from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook
import time

# Initialize Chrome WebDriver
driver = webdriver.Chrome()

# Maximize the window
driver.maximize_window()
driver.implicitly_wait(60)

# Navigate to the login page
wb = load_workbook("ACG_Common_Workbook.xlsx")
#load sheet
typeev = wb["URL_Login_cred_Tenant"]
type_env = typeev.cell(row=2, column=4).value

if type_env == "ACG":
    # fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_ACG"]

if type_env == "Tenant":
    # fetch login_cred sheet
    url_login_cred = wb["URL_Login_cred_Tenant"]

url = url_login_cred.cell(row=2, column=1).value
print(url)
driver.get(str(url))

# Click to accept or dismiss any pop-ups
try:
    WebDriverWait(driver, 3).until(
        EC.element_to_be_clickable((By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[3]/i"))
    ).click()
except Exception as e:
    print(f"Exception during popup handling: {e}")

# Load the login worksheet
login_sheet = wb['Login_user']



# Define expected error messages
expected_errors = {
    "Passwords don't match",
    "User with given creds not available",
    "Fields missing",
    "Cannot read properties of undefined (reading 'close')"
}

# Iterate over rows in the sheet
for row_index, row in enumerate(login_sheet.iter_rows(min_row=2, max_row=login_sheet.max_row, values_only=False), start=2):

    username_cell = row[1]
    password_cell = row[2]
    testcases_cell = row[3]
    cloud_message_cell = row[4]  # This is where the error message will be written
    actual_message_cell = row[5]
    progress_cell = row[6]

    username = username_cell.value or ""
    password = password_cell.value or ""
    testcases = testcases_cell.value or ""

    print(f"Processing User ID: {user_id_cell.value}, Username: '{username}', Password: '{password}'")

    try:
        # Wait and find the username input field
        username_input = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.NAME, "userName"))
        )
        username_input.clear()
        username_input.send_keys(username)

        # Find the password input field
        password_input = WebDriverWait(driver, 10).until(
            EC.visibility_of_element_located((By.XPATH, "//input[@type='password']"))
        )
        password_input.clear()
        password_input.send_keys(password)

        # Find the login button and click it
        login_button = WebDriverWait(driver, 10).until(
            EC.element_to_be_clickable((By.XPATH, "//button[text()='Login']"))
        )
        login_button.click()

        try:
            # Check if there's an error message
            error_element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.CSS_SELECTOR, "div[data-status='error']"))
            )
            error_message = error_element.text
            print(f"Extracted Error Message: {error_message}")

            # Write the error message to the cloud_message_cell
            cloud_message_cell.value = error_message

            # Check if the error message matches any expected errors
            if error_message in expected_errors:
                progress_cell.value = "Fail"
            else:
                progress_cell.value = "Unexpected Error"

            actual_message_cell.value = error_message

        except Exception:
            progress_cell.value = "Pass"  # If no error message, assume pass
            cloud_message_cell.value = "No error message found"  # Write this to cloud_message_cell
            actual_message_cell.value = "No error message found"

    except Exception as e:
        print(f"An error occurred while processing User ID {user_id_cell.value}: {e}")
        progress_cell.value = "Error"
        cloud_message_cell.value = str(e)  # Write the exception message to cloud_message_cell
        actual_message_cell.value = str(e)

    time.sleep(5)  # Adjust sleep time as needed

# Save the updated Excel workbook
try:
    wb.save(file_path)
except Exception as e:
    print(f"An error occurred while saving the Excel file: {e}")

# Close the WebDriver
driver.quit()
