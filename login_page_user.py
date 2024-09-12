import pytest
import http
from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook, load_workbook



def login(chrome_browser):

    # chrome_browser.implicitly_wait(60)
    wb = load_workbook("ACG_Common_Workbook.xlsx")
    typeev = wb["URL_Login_cred_Tenant"]
    type_env = typeev.cell(row=2, column=4).value

    if type_env == "ACG":
        # fetch login_cred sheet
        url_login_cred = wb["URL_Login_cred_ACG"]

    if type_env == "Tenant":
        # fetch login_cred sheet
        url_login_cred = wb["URL_Login_cred_Tenant"]

    url = url_login_cred.cell(row=2, column=1).value
    print(url)
    chrome_browser.get(str(url))

    # if type_env == 'ACG':
    chrome_browser.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[3]/i").click()

    usern = url_login_cred.cell(row=2, column=2).value
    chrome_browser.find_element(By.NAME, 'userName').send_keys(str(usern))

    passw = url_login_cred.cell(row=2, column=3).value
    chrome_browser.find_element(By.NAME, 'password').send_keys(str(passw))

    chrome_browser.find_element(By.XPATH, "//*[@id='root']/div/div/div[4]/div/div[2]/div/button").click()


